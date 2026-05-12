"""
「【助成金申請】書類準備用_会社情報シート.xlsx」から会社情報を取り込むモジュール。

シート構造（2025年度）:
  ①企業基本情報シート C3〜C15 に各項目が入っている
  ②事業所情報シート G22 に常時雇用労働者数の合計
"""
from __future__ import annotations

import re
import warnings
from io import BytesIO
from typing import Optional, Tuple, List

import openpyxl

from .company_form import CompanyInfo, Office


def _extract_postal_address(text: str) -> Tuple[str, str]:
    """'〒000-1111 東京都〇〇区××1-1-1' → ('000-1111', '東京都〇〇区××1-1-1')"""
    if not text:
        return ("", "")
    s = str(text).strip()
    s_norm = s.replace('　', ' ')
    # 〒記号や空白のみの場合は空として扱う
    if re.fullmatch(r'[〒\s]*', s_norm):
        return ("", "")
    m = re.match(r'[〒]?\s*(\d{3})[-−ー―]?(\d{4})\s*(.*)', s_norm)
    if m:
        postal = f"{m.group(1)}-{m.group(2)}"
        address = m.group(3).strip()
        return (postal, address)
    # 郵便番号がなく住所のみの場合
    stripped = re.sub(r'^[〒\s]+', '', s_norm).strip()
    return ("", stripped)


def _split_title_name(text: str) -> Tuple[str, str]:
    """
    代表者役職/氏名文字列を役職と氏名に分割。
    例: '代表取締役 助成金 太郎' → ('代表取締役', '助成金 太郎')
    例: '代表取締役社長 田中 太郎' → ('代表取締役社長', '田中 太郎')
    例: '田中 太郎' → ('代表取締役', '田中 太郎')  役職なければデフォルト
    """
    if not text:
        return ("代表取締役", "")
    s = str(text).strip().replace('　', ' ')
    parts = [p for p in re.split(r'\s+', s) if p]
    if not parts:
        return ("代表取締役", "")
    if len(parts) == 1:
        return ("代表取締役", parts[0])
    if len(parts) == 2:
        # 「田中 太郎」のように姓名の可能性もあるが、助成金書類の慣例では役職+氏名が多い
        # 役職っぽいキーワードを判定
        if any(kw in parts[0] for kw in ('代表', '取締', '社長', '会長', '専務', '常務', '理事', '長')):
            return (parts[0], parts[1])
        # 姓名と判断
        return ("代表取締役", f"{parts[0]} {parts[1]}")
    # 3つ以上: 最後の2つを姓名、残りを役職
    return (' '.join(parts[:-2]), f"{parts[-2]} {parts[-1]}")


def _parse_int(val) -> int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace('人', '').replace(',', '').replace('名', '').strip()
    try:
        return int(s)
    except ValueError:
        return 0


# 47都道府県（住所文字列の先頭マッチ用）
_PREFECTURES = [
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県',
    '岐阜県', '静岡県', '愛知県', '三重県',
    '滋賀県', '京都府', '大阪府', '兵庫県', '奈良県', '和歌山県',
    '鳥取県', '島根県', '岡山県', '広島県', '山口県',
    '徳島県', '香川県', '愛媛県', '高知県',
    '福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県',
    '沖縄県',
]


def detect_prefecture(address: str) -> Optional[str]:
    """住所文字列の先頭から都道府県名を抽出"""
    if not address:
        return None
    s = str(address).strip()
    for pref in _PREFECTURES:
        if s.startswith(pref):
            return pref
    return None


def prefecture_to_labor_bureau(prefecture: Optional[str]) -> str:
    """都道府県名から対応する労働局名を導出
    例: '東京都' → '東京労働局', '北海道' → '北海道労働局',
        '大阪府' → '大阪労働局', '京都府' → '京都労働局'
    """
    if not prefecture:
        return ""
    if prefecture == "北海道":
        return "北海道労働局"
    # 都/府/県 を1文字だけ削除（rstrip だと京都府→京 になるため [:-1] を使う）
    if prefecture[-1] in ("都", "府", "県"):
        return prefecture[:-1] + "労働局"
    return prefecture + "労働局"


def _find_employee_count_by_label(ws, max_row: int = 30) -> int:
    """シート内で「従業員数」「常時雇用」を含むラベルを探し、隣接セルの値を返す

    ラベル検索の優先度: 同じ行の右側 → 次の列
    """
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value is None:
                continue
            label = str(cell.value).strip()
            if not any(kw in label for kw in ('従業員数', '常時雇用')):
                continue
            # 同じ行の右側を探す
            for col_offset in range(1, 12):
                target = ws.cell(row=cell.row, column=cell.column + col_offset)
                emp = _parse_int(target.value)
                if emp > 0:
                    return emp
    return 0


def _find_sheet(wb, *keywords) -> Optional[str]:
    """シート名に指定キーワードが含まれる最初のシート名を返す"""
    for sn in wb.sheetnames:
        if any(kw in sn for kw in keywords):
            return sn
    return None


def _get(ws, coord: str) -> str:
    v = ws[coord].value
    return str(v).strip() if v is not None else ""


def _is_example_value(s: str) -> bool:
    """「例）...」「○○○○」等のサンプル値を判定"""
    if not s:
        return True
    return ('例）' in s or '例)' in s or '○○○○' in s or '〇〇〇〇' in s)


def _build_insurance_number(p1: str, p2: str, p3: str) -> str:
    """4桁-6桁-1桁 を組み立て（いずれかが空ならそれを空のまま）"""
    if not p1 and not p2 and not p3:
        return ""
    return f"{p1}-{p2}-{p3}"


def _extract_offices(wb) -> List[Office]:
    """②事業所情報シートから事業所リストを抽出"""
    offices: List[Office] = []
    sheet_name = _find_sheet(wb, '事業所情報', '②')
    if not sheet_name:
        return offices
    ws = wb[sheet_name]

    # 申請事業所 (行8)
    name8 = _get(ws, 'B8')
    if name8 and not _is_example_value(name8):
        c8 = _get(ws, 'C8')
        d8 = _get(ws, 'D8')
        e8 = _get(ws, 'E8')
        emp8 = _parse_int(ws['F8'].value)
        offices.append(Office(
            name=name8,
            insurance_number=_build_insurance_number(c8, d8, e8),
            employee_count=emp8,
        ))

    # 従たる事業所 (行13〜21の最大9件)
    for row in range(13, 22):
        name = _get(ws, f'B{row}')
        if not name or _is_example_value(name):
            continue
        c = _get(ws, f'C{row}')
        d = _get(ws, f'D{row}')
        e = _get(ws, f'E{row}')
        emp = _parse_int(ws[f'F{row}'].value)
        offices.append(Office(
            name=name,
            insurance_number=_build_insurance_number(c, d, e),
            employee_count=emp,
        ))

    return offices


def import_company_from_excel(file_bytes: bytes) -> Tuple[Optional[CompanyInfo], List[str]]:
    """
    会社情報シートExcelから CompanyInfo を復元。

    Returns:
        (CompanyInfo | None, warnings: list[str])
    """
    import_warnings: List[str] = []
    try:
        # openpyxl の Data Validation 警告など無害な UserWarning を抑制
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return (None, [f"Excelファイルを読み込めません: {e}"])

    base_sheet_name = _find_sheet(wb, '企業基本情報', '①')
    if not base_sheet_name:
        return (None, ["①企業基本情報シートが見つかりません"])
    ws = wb[base_sheet_name]

    company_name = _get(ws, 'C3')
    rep_title, rep_name = _split_title_name(_get(ws, 'C4'))
    postal_code, address = _extract_postal_address(_get(ws, 'C5'))
    head_phone = _get(ws, 'C6')
    corporate_number = _get(ws, 'C8').replace('-', '').replace('−', '').replace('ー', '').strip()
    employee_count_main = _parse_int(ws['C9'].value)
    insurance_office_number = _get(ws, 'C10')
    contact_name = _get(ws, 'C14')
    contact_phone = _get(ws, 'C15')

    # 雇用保険適用事業所番号の形式を整える (4桁-6桁-1桁)
    if insurance_office_number:
        ion = re.sub(r'[−ー―\s]', '-', insurance_office_number)
        digits = re.sub(r'[^0-9]', '', ion)
        if len(digits) == 11:
            insurance_office_number = f"{digits[:4]}-{digits[4:10]}-{digits[10]}"
        elif '-' in ion:
            insurance_office_number = ion

    # ②事業所情報シートから労働者数合計を取得
    employee_count_total = 0
    office_sheet_name = _find_sheet(wb, '事業所情報', '②')
    if office_sheet_name:
        ws2 = wb[office_sheet_name]
        # G22 or C22 いずれかに合計があるはず
        employee_count_total = _parse_int(ws2['G22'].value) or _parse_int(ws2['C22'].value)
        # それでも取れない場合はラベル検索
        if employee_count_total == 0:
            employee_count_total = _find_employee_count_by_label(ws2)

    # ①シートでも C9 で取れなければラベル検索を試みる
    if employee_count_main == 0:
        employee_count_main = _find_employee_count_by_label(ws)

    employee_count = employee_count_total if employee_count_total > 0 else employee_count_main
    if employee_count == 0:
        employee_count = 1  # dataclassのデフォルトとの互換性
        import_warnings.append("従業員数が取得できませんでした。手動で入力してください。")

    phone_number = contact_phone or head_phone

    offices = _extract_offices(wb)

    # 本社（申請事業所）の情報が取れていれば、CompanyInfo の番号・名称と整合させる
    if offices:
        head = offices[0]
        if not insurance_office_number and head.insurance_number:
            insurance_office_number = head.insurance_number

    # 本社所在地の都道府県から労働局を自動判定
    pref = detect_prefecture(address)
    labor_bureau = prefecture_to_labor_bureau(pref) if pref else ""

    company = CompanyInfo(
        company_name=company_name,
        insurance_office_number=insurance_office_number,
        postal_code=postal_code,
        address=address,
        contact_name=contact_name,
        contact_department="",
        contact_email="",
        phone_number=phone_number,
        main_business="",  # このシートには含まれない（産業分類番号のみ）
        employee_count=employee_count,
        labor_bureau=labor_bureau,
        representative_name=rep_name,
        representative_title=rep_title,
        corporate_number=corporate_number,
        offices=offices,
    )

    # 不足項目の警告
    if not company_name:
        import_warnings.append("企業名（C3）が空欄です")
    if not address:
        import_warnings.append("本社所在地（C5）が空欄または郵便番号形式が不正です")
    if not contact_name:
        import_warnings.append("助成金担当者名（C14）が空欄です")
    if labor_bureau:
        import_warnings.append(f"管轄労働局を住所から自動判定しました：{labor_bureau}")
    else:
        import_warnings.append("「管轄労働局」は住所から自動判定できなかったため、手動で選択してください。")
    import_warnings.append("「主たる事業」はシートに含まれないため、手動で入力してください。")

    return (company, import_warnings)
