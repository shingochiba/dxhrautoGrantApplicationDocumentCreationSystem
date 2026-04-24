"""受講者一覧入力テンプレートを作成するスクリプト"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_participant_template():
    """受講者一覧入力テンプレートを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受講者一覧"

    # スタイル設定
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ヘッダー行の設定
    headers = [
        ("No", 5),
        ("氏名", 20),
        ("雇用保険被保険者番号\n(4桁-6桁-1桁)", 25),
        ("雇用形態\n(正規雇用/有期契約)", 18),
        ("カリキュラム名", 30),
        ("訓練開始日\n(YYYY/MM/DD)", 15),
        ("訓練終了日\n(YYYY/MM/DD)", 15),
        ("受講料(税抜)", 15),
        ("助成コース\n(リスキリング/定額制)", 20),
    ]

    # ヘッダーを書き込み
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # サンプルデータ行（2-11行目）
    sample_data = [
        (1, "山田 太郎", "1234-567890-1", "正規雇用", "DX基礎研修", "2026/05/01", "2026/07/31", 50000, "リスキリング"),
        (2, "佐藤 花子", "2345-678901-2", "正規雇用", "DX基礎研修", "2026/05/01", "2026/07/31", 50000, "リスキリング"),
        (3, "鈴木 一郎", "3456-789012-3", "有期契約", "AI活用研修", "2026/06/01", "2026/08/31", 80000, "定額制"),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = center_align

    # 空のデータ行を追加（4-100行目）
    for row_idx in range(5, 101):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 1:
                cell.value = row_idx - 1
                cell.alignment = center_align

    # 入力説明シートを追加
    ws2 = wb.create_sheet(title="入力説明")
    instructions = [
        ["【受講者一覧 入力説明】", ""],
        ["", ""],
        ["項目", "説明"],
        ["No", "自動採番されます（変更不要）"],
        ["氏名", "受講者の氏名を入力（例：山田 太郎）"],
        ["雇用保険被保険者番号", "11桁のハイフン区切り形式で入力（例：1234-567890-1）"],
        ["雇用形態", "「正規雇用」または「有期契約」を選択"],
        ["カリキュラム名", "受講する研修・カリキュラムの名称"],
        ["訓練開始日", "YYYY/MM/DD形式で入力（例：2026/05/01）"],
        ["訓練終了日", "YYYY/MM/DD形式で入力（例：2026/07/31）"],
        ["受講料(税抜)", "税抜金額を数値で入力"],
        ["助成コース", "「リスキリング」または「定額制」を選択"],
        ["", ""],
        ["【助成コースについて】", ""],
        ["リスキリング", "事業展開等リスキリング支援コース - DX化・グリーン化に伴う訓練向け"],
        ["定額制", "人への投資促進コース（定額制訓練） - サブスク型e-ラーニング向け"],
    ]

    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 3:
                cell.font = header_font
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 60

    # 保存
    output_path = "templates/受講者一覧_入力テンプレート.xlsx"
    wb.save(output_path)
    print(f"テンプレートを作成しました: {output_path}")
    return output_path

if __name__ == "__main__":
    create_participant_template()
